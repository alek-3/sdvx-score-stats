import csv
import os
import openpyxl
import datetime

from avarage_scores import AverageScores


class Stats:
    def output_score_stats(self, file_name):
        csv_path = os.path.join("scoresheet", file_name)
        self.show_play_count(csv_path)

        average_scores = AverageScores(file_name)
        self.write_excel(average_scores.record_date, average_scores.averages)

    def show_play_count(self, csv_path):
        """これまでプレーした回数を譜面単位で（投入クレジット数ではない）を出力する"""
        with open(csv_path, "r", encoding="utf-8") as f:
            r = csv.DictReader(f, delimiter=",")
            play_count = 0
            for row in r:
                play_count += int(row["プレー回数"])
            print("譜面プレー回数：{play_count}".format(play_count=play_count))

    def write_excel(self, record_date, averages):
        """日付と平均スコアの推移をExcelシートに書きこむ"""
        wb = openpyxl.load_workbook('sdvx_score_averages.xlsx')
        sheet = wb.active

        # 1列目は日付を入力する
        sheet.cell(row=file_number + 2, column=1).value = record_date

        # 3行目以降
        # 2列目以降
        level = 1
        for average in averages:
            sheet.cell(row=file_number + 2, column=level + 1).value = average[1]['平均スコア']
            sheet.cell(row=2, column=level + 1).value = average[1]['譜面数']
            level += 1

        wb.save('sdvx_score_averages.xlsx')

    def write_header(self):
        if os.path.exists('sdvx_score_averages.xlsx'):
            wb = openpyxl.load_workbook('sdvx_score_averages.xlsx')
        else:
            wb = openpyxl.Workbook()

        sheet = wb.active
        sheet.cell(row=1, column=1).value = '日付＼レベル'
        sheet.cell(row=2, column=1).value = '譜面数'
        for col in range(2,22):
            sheet.cell(row=1, column=col).value = col - 1

        wb.save('sdvx_score_averages.xlsx')

    def arrange_file_layout(self):
        wb = openpyxl.load_workbook('sdvx_score_averages.xlsx')

        sheet = wb.active
        sheet.column_dimensions['A'].width = 13
        sheet.column_dimensions.group(start='B', end='M', hidden=True)

        wb.save('sdvx_score_averages.xlsx')


stats = Stats()
files = os.listdir("scoresheet")
files_file = [f for f in files if os.path.isfile(os.path.join("scoresheet", f))]

stats.write_header()

# ファイル番号を行番号として使う
file_number = 1

for file in files_file:
    stats.output_score_stats(file)
    file_number += 1

stats.arrange_file_layout()
